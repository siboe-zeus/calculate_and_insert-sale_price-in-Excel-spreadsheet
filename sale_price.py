import openpyxl
from decimal import Decimal, ROUND_UP


def sale_price(filename):
    wb = openpyxl.load_workbook(filename)
    sheet = wb["SalesOrders"]
    sheet["h1"] = "Unit Sale Price"
    for row in range(2, sheet.max_row + 1):
         cell = sheet.cell(row, 6)
         unit_sale_price = float(cell.value) * 1.05
         unit_sale_price_cell = sheet.cell(row, 8)
         unit_sale_price_cell.value = Decimal(round(unit_sale_price, 2))
         (unit_sale_price_cell.value * 2).quantize(Decimal(".1"), rounding=ROUND_UP) / 2

    wb.save(filename)


sale_price(r"C:\Users\kenne\Desktop\SampleData.xlsx")
